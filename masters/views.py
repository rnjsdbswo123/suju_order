from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import TemplateView, View, ListView, UpdateView
from django.urls import reverse_lazy
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from django.db import transaction
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.db.models import Q
import openpyxl
from SujuOrderSystem.utils import FACILITY_LIST
from django.core.paginator import Paginator
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required
from users.permissions import is_in_role

# 모델 가져오기
from .models import Customer, Product, CustomerProductMap, SalesFavoriteProduct, RawMaterial

# ==========================================
# 1. [화면] 엑셀 데이터 일괄 업로드
# ==========================================
class DataUploadView(LoginRequiredMixin, TemplateView):
    template_name = 'masters/data_upload.html'

    def get(self, request, *args, **kwargs):
        return render(request, self.template_name)

    def post(self, request):
        try:
            if 'customer_file' in request.FILES:
                self.upload_customers(request.FILES['customer_file'])
                messages.success(request, "거래처 업로드 완료! 🎉")
            
            elif 'product_file' in request.FILES:
                self.upload_products(request.FILES['product_file'])
                messages.success(request, "품목 업로드 완료! 🎉")
                
            elif 'mapping_file' in request.FILES:
                self.upload_mappings(request.FILES['mapping_file'])
                messages.success(request, "매핑 업로드 완료! 🎉")

            elif 'rawmaterial_file' in request.FILES:
                self.upload_rawmaterials(request.FILES['rawmaterial_file'])
                messages.success(request, "부자재 업로드 완료! 🎉")

        except Exception as e:
            messages.error(request, f"업로드 중 오류 발생: {str(e)}")
        return redirect('data-upload')

    @transaction.atomic
    def upload_rawmaterials(self, file):
        wb = openpyxl.load_workbook(file)
        ws = wb.active

        materials_data = {}
        # 엑셀 헤더: 부자재명, 부자재코드, 바코드, 단가, 활성여부
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 2 or not row[1]:
                continue
            
            sku = str(row[1]).strip()
            if not sku:
                continue

            name = str(row[0]).strip()
            barcode = str(row[2]).strip() if len(row) > 2 and row[2] else None
            price = row[3] if len(row) > 3 and row[3] is not None else 0

            is_active = None
            if len(row) > 4:
                excel_val = row[4]
                if excel_val is False or (isinstance(excel_val, str) and excel_val.strip().upper() == 'FALSE'):
                    is_active = False
                elif excel_val is not None:
                    is_active = True
            
            materials_data[sku] = {
                'name': name, 'barcode': barcode, 'unit_price': price, 'is_active': is_active
            }

        existing_materials = RawMaterial.objects.filter(sku__in=materials_data.keys())
        existing_materials_map = {m.sku: m for m in existing_materials}

        materials_to_create = []
        materials_to_update = []
        update_fields = set()

        for sku, data in materials_data.items():
            if sku in existing_materials_map:
                material = existing_materials_map[sku]
                should_update = False
                
                if material.name != data['name']:
                    material.name = data['name']
                    should_update = True
                    update_fields.add('name')
                
                if data['barcode'] is not None and material.barcode != data['barcode']:
                    material.barcode = data['barcode']
                    should_update = True
                    update_fields.add('barcode')

                if material.unit_price != data['unit_price']:
                    material.unit_price = data['unit_price']
                    should_update = True
                    update_fields.add('unit_price')

                if data['is_active'] is not None and material.is_active != data['is_active']:
                    material.is_active = data['is_active']
                    should_update = True
                    update_fields.add('is_active')
                
                if should_update:
                    materials_to_update.append(material)
            else:
                materials_to_create.append(RawMaterial(
                    sku=sku,
                    name=data['name'],
                    barcode=data['barcode'],
                    unit_price=data['unit_price'],
                    is_active=data['is_active'] if data['is_active'] is not None else True
                ))
        
        if materials_to_create:
            RawMaterial.objects.bulk_create(materials_to_create)
        
        if materials_to_update:
            RawMaterial.objects.bulk_update(materials_to_update, list(update_fields))


    @transaction.atomic
    def upload_customers(self, file):
        wb = openpyxl.load_workbook(file)
        ws = wb.active

        customers_data = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            
            name = row[0].strip()
            if not name:
                continue

            biz_id = row[1] if len(row) > 1 and row[1] else None
            
            is_active = None
            if len(row) > 2:
                excel_val = row[2]
                if excel_val is False or (isinstance(excel_val, str) and excel_val.strip().upper() == 'FALSE'):
                    is_active = False
                elif excel_val is not None:
                    is_active = True
            
            customers_data[name] = {'business_id': biz_id, 'is_active': is_active}

        existing_customers = Customer.objects.filter(name__in=customers_data.keys())
        existing_customers_map = {c.name: c for c in existing_customers}

        customers_to_create = []
        customers_to_update = []
        update_fields = set()

        for name, data in customers_data.items():
            biz_id = data['business_id']
            is_active = data['is_active']

            if name in existing_customers_map:
                customer = existing_customers_map[name]
                
                # Check for updates
                should_update = False
                if biz_id is not None and customer.business_id != biz_id:
                    customer.business_id = biz_id
                    should_update = True
                    update_fields.add('business_id')

                if is_active is not None and customer.is_active != is_active:
                    customer.is_active = is_active
                    should_update = True
                    update_fields.add('is_active')
                
                if should_update:
                    customers_to_update.append(customer)
            else:
                customers_to_create.append(Customer(name=name, business_id=biz_id, is_active=is_active if is_active is not None else True))
        
        if customers_to_create:
            Customer.objects.bulk_create(customers_to_create)
        
        if customers_to_update:
            Customer.objects.bulk_update(customers_to_update, list(update_fields))

    @transaction.atomic
    def upload_products(self, file):
        wb = openpyxl.load_workbook(file)
        ws = wb.active

        products_data = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 2 or not row[1]:
                continue
            
            sku = row[1].strip()
            if not sku:
                continue

            name = row[0] if row[0] else ''
            price = row[2] if len(row) > 2 and row[2] is not None else 0
            facility = row[3] if len(row) > 3 and row[3] else 'A동'
            
            products_data[sku] = {'name': name, 'unit_price': price, 'production_facility': facility}

        existing_products = Product.objects.filter(sku__in=products_data.keys())
        existing_products_map = {p.sku: p for p in existing_products}

        products_to_create = []
        products_to_update = []
        update_fields = {'name', 'unit_price', 'production_facility'}

        for sku, data in products_data.items():
            if sku in existing_products_map:
                product = existing_products_map[sku]
                
                # Check for updates
                if (product.name != data['name'] or 
                        product.unit_price != data['unit_price'] or
                        product.production_facility != data['production_facility']):
                    
                    product.name = data['name']
                    product.unit_price = data['unit_price']
                    product.production_facility = data['production_facility']
                    products_to_update.append(product)
            else:
                products_to_create.append(Product(
                    sku=sku,
                    name=data['name'],
                    unit_price=data['unit_price'],
                    production_facility=data['production_facility']
                ))
        
        if products_to_create:
            Product.objects.bulk_create(products_to_create)
        
        if products_to_update:
            Product.objects.bulk_update(products_to_update, list(update_fields))

    @transaction.atomic
    def upload_mappings(self, file):
        wb = openpyxl.load_workbook(file)
        ws = wb.active

        mappings_to_process = set()
        customer_names = set()
        product_skus = set()

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 2 or not row[0] or not row[1]:
                continue
            
            c_name = row[0].strip()
            p_sku = row[1].strip()

            if c_name and p_sku:
                mappings_to_process.add((c_name, p_sku))
                customer_names.add(c_name)
                product_skus.add(p_sku)

        # 한 번의 쿼리로 필요한 모든 고객과 제품을 가져옵니다.
        customers = Customer.objects.filter(name__in=customer_names)
        products = Product.objects.filter(sku__in=product_skus)

        # 빠른 조회를 위해 딕셔너리로 변환합니다.
        customer_map = {c.name: c.id for c in customers}
        product_map = {p.sku: p.id for p in products}

        mappings_to_create = []
        for c_name, p_sku in mappings_to_process:
            customer_id = customer_map.get(c_name)
            product_id = product_map.get(p_sku)

            if customer_id and product_id:
                mappings_to_create.append(
                    CustomerProductMap(customer_id=customer_id, product_id=product_id)
                )
        
        if mappings_to_create:
            CustomerProductMap.objects.bulk_create(mappings_to_create, ignore_conflicts=True)

# ==========================================
# 2. [화면] 거래처-품목 매핑 직접 관리
# ==========================================
class CustomerProductManageView(LoginRequiredMixin, TemplateView):
    template_name = 'masters/customer_product_list.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['mappings'] = CustomerProductMap.objects.select_related('customer', 'product').order_by('-id')
        return context

    def post(self, request):
        customer_id = request.POST.get('customer')
        product_id = request.POST.get('product')
        if customer_id and product_id:
            if not CustomerProductMap.objects.filter(customer_id=customer_id, product_id=product_id).exists():
                CustomerProductMap.objects.create(customer_id=customer_id, product_id=product_id)
        return redirect('customer-product-manage')

# ==========================================
# 2.1. [화면] 거래처-품목 매핑 직접 관리 (개선)
# ==========================================
class CustomerProductMappingView(LoginRequiredMixin, TemplateView):
    template_name = 'masters/customer_product_mapping.html'

def delete_customer_product(request, pk):
    mapping = get_object_or_404(CustomerProductMap, pk=pk)
    mapping.delete()
    return redirect('customer-product-manage')

# ==========================================
# 3. [API] 데이터 검색 및 조회 (AJAX용)
# ==========================================
@api_view(['GET'])
def get_products_by_customer(request, customer_id):
    mappings = CustomerProductMap.objects.filter(customer_id=customer_id).select_related('product').order_by('product__order', 'product__name')
    data = [{"id": m.product.id, "name": m.product.name, "sku": m.product.sku, "price": m.product.unit_price} for m in mappings]
    return Response(data)

@api_view(['GET'])
def search_customers(request):
    query = request.GET.get('q', '')
    if query:
        customers = Customer.objects.filter(
            Q(name__icontains=query) | Q(business_id__icontains=query)
        ).filter(is_active=True)
    else:
        customers = Customer.objects.filter(is_active=True)
    customers = customers[:20]
    data = [{"id": c.id, "text": c.name} for c in customers]
    return Response({"results": data})

@api_view(['GET'])
def customer_detail(request, customer_id):
    customer = get_object_or_404(Customer, id=customer_id)
    return Response({"id": customer.id, "text": customer.name})

@api_view(['GET'])
def search_products(request):
    query = request.GET.get('q', '')
    if query:
        products = Product.objects.filter(
            Q(name__icontains=query) | Q(sku__icontains=query)
        )
    else:
        # 개선된 UI에서는 모든 품목을 가져와야 하므로 all() 사용. 페이지네이션 등 성능 고려 필요.
        products = Product.objects.all().order_by('name')

    # 모든 품목을 반환하도록 수정 (기존 :20 제한 해제)
    data = [
        {"id": p.id, "text": f"{p.name} ({p.sku})"} 
        for p in products
    ]
    return Response({"results": data})

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def update_customer_products(request):
    customer_id = request.data.get('customer_id')
    product_ids = request.data.get('product_ids', [])

    if not customer_id:
        return Response({"error": "거래처 ID가 필요합니다."}, status=400)

    try:
        with transaction.atomic():
            # 기존 매핑 삭제
            CustomerProductMap.objects.filter(customer_id=customer_id).delete()
            
            # 새 매핑 추가
            new_mappings = [
                CustomerProductMap(customer_id=customer_id, product_id=pid)
                for pid in product_ids
            ]
            CustomerProductMap.objects.bulk_create(new_mappings)
            
        return Response({"message": f"{len(product_ids)}개의 품목 매핑을 저장했습니다."})

    except Exception as e:
        return Response({"error": f"저장 중 오류 발생: {str(e)}"}, status=500)


# ==========================================
# 4. [화면] 영업사원 선호품목 관리
# ==========================================
class SalesFavoriteProductManageView(LoginRequiredMixin, View):
    template_name = 'masters/sales_favorite_product_manage.html'

    def get(self, request, *args, **kwargs):
        all_products = Product.objects.filter(is_active=True).order_by('name')
        favorite_product_ids = SalesFavoriteProduct.objects.filter(user=request.user).values_list('product_id', flat=True)
        context = {
            'products': all_products,
            'favorite_product_ids': set(favorite_product_ids),
            'page_title': '내 선호품목 관리'
        }
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        product_id = request.POST.get('product_id')
        action = request.POST.get('action')
        if not product_id or not action:
            messages.error(request, '잘못된 요청입니다.')
            return redirect('sales-favorite-manage')
        product = get_object_or_404(Product, id=product_id)
        if action == 'add':
            if not SalesFavoriteProduct.objects.filter(user=request.user, product=product).exists():
                SalesFavoriteProduct.objects.create(user=request.user, product=product)
                messages.success(request, f'"{product.name}"을(를) 선호 품목에 추가했습니다.')
        elif action == 'remove':
            favorite = SalesFavoriteProduct.objects.filter(user=request.user, product=product)
            if favorite.exists():
                favorite.delete()
                messages.success(request, f'"{product.name}"을(를) 선호 품목에서 제거했습니다.')
        return redirect('sales-favorite-manage')

# ==========================================
# 5. [화면] 품목별 생산동 관리
# ==========================================
class ProductFacilityManageView(LoginRequiredMixin, View):
    template_name = 'product_facility_manage.html'

    def get(self, request, *args, **kwargs):
        product_list = Product.objects.all().order_by('name')
        
        # 검색
        q = request.GET.get('q', '')
        if q:
            product_list = product_list.filter(Q(name__icontains=q) | Q(sku__icontains=q))

        # 생산동 필터
        facility = request.GET.get('facility', 'ALL')
        if facility == '미지정':
            product_list = product_list.filter(Q(production_facility__isnull=True) | Q(production_facility=''))
        elif facility != 'ALL':
            product_list = product_list.filter(production_facility=facility)

        # 페이지네이션
        paginator = Paginator(product_list, 25)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        
        context = {
            'page_title': '품목별 생산동 관리',
            'page_obj': page_obj,
            'facility_list': FACILITY_LIST,
            'search_query': q,
            'selected_facility': facility,
        }
        return render(request, self.template_name, context)

    @transaction.atomic
    def post(self, request, *args, **kwargs):
        # 엑셀 파일 업로드 처리
        if 'facility_file' in request.FILES:
            try:
                file = request.FILES['facility_file']
                wb = openpyxl.load_workbook(file)
                ws = wb.active
                
                facility_data = {}
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or len(row) < 2 or not row[0]:
                        continue
                    sku = str(row[0]).strip()
                    facility = str(row[1]).strip()
                    if sku:
                        facility_data[sku] = facility

                products_to_update = []
                # 파일에 있는 SKU에 해당하는 모든 제품을 한 번에 가져옵니다.
                products = Product.objects.filter(sku__in=facility_data.keys())

                for product in products:
                    new_facility = facility_data.get(product.sku)
                    # 생산동 정보가 변경된 경우에만 업데이트 목록에 추가합니다.
                    if new_facility is not None and product.production_facility != new_facility:
                        product.production_facility = new_facility
                        products_to_update.append(product)
                
                if products_to_update:
                    Product.objects.bulk_update(products_to_update, ['production_facility'])
                    messages.success(request, f"{len(products_to_update)}개 품목의 생산동 정보가 엑셀로 업데이트되었습니다.")
                else:
                    messages.info(request, "업데이트할 내용이 없습니다.")

            except Exception as e:
                messages.error(request, f"엑셀 처리 중 오류 발생: {e}")

        # 개별 업데이트 처리
        elif 'update_individual' in request.POST:
            product_ids = request.POST.getlist('product_id')
            products_to_update = []
            
            # 개별 업데이트 대상 제품들을 한 번에 가져옵니다.
            products = Product.objects.in_bulk(product_ids)
            
            for pid in product_ids:
                product = products.get(int(pid))
                if product:
                    new_facility = request.POST.get(f'production_facility_{pid}')
                    if product.production_facility != new_facility:
                        product.production_facility = new_facility
                        products_to_update.append(product)
            
            if products_to_update:
                Product.objects.bulk_update(products_to_update, ['production_facility'])
                messages.success(request, f"{len(products_to_update)}개 품목의 생산동 정보가 업데이트되었습니다.")

        return redirect('product-facility-manage')

# ==========================================
# 6. [화면] 신규 거래처/품목 개별 등록
# ==========================================
class CustomerCreateView(LoginRequiredMixin, View):
    def get(self, request):
        return render(request, 'masters/customer_form.html')

    def post(self, request):
        name = request.POST.get('name')
        business_id = request.POST.get('business_id')

        if not name:
            messages.error(request, "거래처명은 필수입니다.")
            return render(request, 'masters/customer_form.html', {'name': name, 'business_id': business_id})

        if Customer.objects.filter(name=name).exists():
            messages.error(request, "이미 존재하는 거래처명입니다.")
            return render(request, 'masters/customer_form.html', {'name': name, 'business_id': business_id})

        Customer.objects.create(name=name, business_id=business_id)
        messages.success(request, f"거래처 '{name}'이(가) 성공적으로 등록되었습니다.")
        return redirect('data-upload')

class ProductCreateView(LoginRequiredMixin, View):
    def get(self, request):
        context = {'facility_list': FACILITY_LIST}
        return render(request, 'masters/product_form.html', context)

    def post(self, request):
        name = request.POST.get('name')
        sku = request.POST.get('sku')
        unit_price = request.POST.get('unit_price', 0)
        production_facility = request.POST.get('production_facility')

        form_data = {
            'name': name, 'sku': sku, 'unit_price': unit_price, 
            'production_facility': production_facility
        }
        context = {'form_data': form_data, 'facility_list': FACILITY_LIST}

        if not name or not sku:
            messages.error(request, "품목명과 품목코드는 필수입니다.")
            return render(request, 'masters/product_form.html', context)

        if Product.objects.filter(sku=sku).exists():
            messages.error(request, "이미 존재하는 품목코드입니다.")
            return render(request, 'masters/product_form.html', context)

        Product.objects.create(
            name=name,
            sku=sku,
            unit_price=unit_price,
            production_facility=production_facility
        )
        messages.success(request, f"품목 '{name}'이(가) 성공적으로 등록되었습니다.")
        return redirect('data-upload')


class RawMaterialCreateView(LoginRequiredMixin, View):
    def get(self, request):
        return render(request, 'masters/rawmaterial_form.html')

    def post(self, request):
        name = request.POST.get('name')
        sku = request.POST.get('sku')
        barcode = request.POST.get('barcode')
        unit_price = request.POST.get('unit_price', 0)
        image = request.FILES.get('image')

        form_data = {
            'name': name, 'sku': sku, 'barcode': barcode, 'unit_price': unit_price
        }
        context = {'form_data': form_data}

        if not name or not sku:
            messages.error(request, "부자재명과 부자재코드는 필수입니다.")
            return render(request, 'masters/rawmaterial_form.html', context)

        if RawMaterial.objects.filter(sku=sku).exists():
            messages.error(request, "이미 존재하는 부자재코드입니다.")
            return render(request, 'masters/rawmaterial_form.html', context)

        RawMaterial.objects.create(
            name=name,
            sku=sku,
            barcode=barcode,
            unit_price=unit_price,
            image=image
        )
        messages.success(request, f"부자재 '{name}'이(가) 성공적으로 등록되었습니다.")
        return redirect('data-upload')


class RawMaterialListView(LoginRequiredMixin, ListView):
    model = RawMaterial
    template_name = 'masters/rawmaterial_list.html'
    context_object_name = 'materials'
    paginate_by = 20
    ordering = ['name']


class RawMaterialUpdateView(LoginRequiredMixin, UpdateView):
    model = RawMaterial
    fields = ['name', 'sku', 'barcode', 'unit_price', 'image', 'is_active']
    template_name = 'masters/rawmaterial_form.html'
    success_url = reverse_lazy('rawmaterial-list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['is_edit_mode'] = True
        return context

@require_POST
@login_required
def rawmaterial_delete(request, pk):
    # '관리자' 또는 '자재담당자' 역할이 있는지 확인
    if not (is_in_role(request.user, '관리자') or is_in_role(request.user, '자재담당자')):
        messages.error(request, '삭제 권한이 없습니다.')
        return redirect('rawmaterial-list')

    material = get_object_or_404(RawMaterial, pk=pk)
    try:
        material.delete()
        messages.success(request, f"부자재 '{material.name}'이(가) 성공적으로 삭제되었습니다.")
    except Exception as e:
        # ForeignKey 제약 조건 등 DB 레벨에서 발생하는 예외 처리
        messages.error(request, f"삭제 중 오류가 발생했습니다: 이 부자재를 사용하고 있는 다른 데이터가 있어 삭제할 수 없습니다.")
    
    return redirect('rawmaterial-list')
